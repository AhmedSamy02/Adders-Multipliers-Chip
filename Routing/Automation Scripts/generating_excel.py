import re
import openpyxl
import numpy as np
import os
import pandas as pd

def extract_min_timing_slack(report):
    slack_match = re.search(r'slack \(MET\)\s+([\d.-]+)', report)
    return float(slack_match.group(1)) if slack_match else None

def extract_timing_slack(report):
    slack_match = re.search(r'slack \(MET\)\s+([\d.-]+)', report)
    return float(slack_match.group(1)) if slack_match else None

def extract_clock(report):
    clock_match = re.search(r'Critical Path Clk Period:\s+([\d.]+)', report)
    return float(clock_match.group(1)) if clock_match else None

def extract_power(report):
    total_power_match = re.search(r'nW \s+([\d.]+)', report)
    leakage_power_match = re.search(r'Cell Leakage Power     = \s+([\d.]+)', report)
    arr = []
    if total_power_match:
        arr.append(float(total_power_match.group(1)))
    if leakage_power_match:
        arr.append(float(leakage_power_match.group(1)))
    return arr if arr else None

def extract_area_and_latches(report):
    area_match = re.search(r'Total cell area:\s+([\d.]+)', report)
    latches_match = re.search(r'Number of sequential cells:\s+([\d.]+)', report)
    arr = []
    if area_match:
        arr.append(float(area_match.group(1)))
    if latches_match:
        arr.append(float(latches_match.group(1)))
    return arr if arr else None

def process_design(index, design, sheet):
    # Specify the path to your report file
    report_file_path = "Outputs/Multipliers/" + design + "/"
    design = design[2:]
    try:
        # Read the report from the file
        with open(report_file_path + design + '_synth_timing.rpt', 'r') as report_file:
            report_string = report_file.read()
        # Extract slack value from the report
        slack_value = extract_min_timing_slack(report_string)
        if slack_value is not None:
            # Add slack value to Excel sheet
            df[index+1][1] = slack_value
            print(df)
            print(f"Max Delay value added for {design}")
        else:
            print(f"Error during extracting timing for {design}")

        # Read the report from the file
        with open(report_file_path + design + '_synth_min_timing.rpt', 'r') as report_file:
            report_string = report_file.read()
        # Extract slack value from the report
        slack_value = extract_timing_slack(report_string)
        if slack_value is not None:
            # Add slack value to Excel sheet
            df[index+1] [2]= slack_value
            print(f"Min Delay value added for {design}")
        else:
            print(f"Error during extracting timing for {design}")

        with open(report_file_path + design + '_synth_qor.rpt', 'r') as report_file:
            report_string = report_file.read()
        
        clk = extract_clock(report_string)
        if clk is not None:
            df[index+1][3] = clk
            print(f"Clock value added for {design}")
        else:
            print(f"Error during extracting clock for {design}")

        with open(report_file_path + design + '_synth_power.rpt', 'r') as report_file:
            report_string = report_file.read()
        total_power,leakage_power = extract_power(report_string)

        if total_power is not None:
            df[index+1][4] = str(total_power) + "nW"
            print(f"Total Power value added for {design}")
        else:
            print(f"Error during extracting total power for {design}")

        if leakage_power is not None:
            df[index+1][6] = str(round((1-(leakage_power/total_power) ) * 100 , ndigits=3)) + "%"
            print(f"Leakage Power value added for {design}")
        else:
            print(f"Error during extracting leakage power for {design}")

        with open(report_file_path + design + '_synth_area.rpt', 'r') as report_file:
            report_string = report_file.read()

        area_match = extract_area_and_latches(report_string)
        if area_match:
            df[index+1][5] = area_match[0]
            print(f"Area value added for {design}")
            df[index+1][7] = area_match[1]
            print(f"# of Latches value added for {design}")
        else:
            print(f"Error during extracting area for {design}")
            print(f"Error during extracting # of Latches for {design}")

           
    except FileNotFoundError:
        print(f"Error: Report files not found for {design}")

if __name__ == "__main__":
    arr = [
    ['', 'Float Multiplier', 'Multiplier Tree', 'Original Booth', 'Sequential Multiplier', 'Verilog (*)'],
    ['Min Delay', '', '', '', '', ''],
    ['Max Delay', '', '', '', '', ''],
    ['Clock', '', '', '', '', ''],
    ['Total Power', '', '', '', '', ''],
    ['Area', '', '', '', '', ''],
    ['Utilization', '', '', '', '', ''],
    ['# of Latches', '', '', '', '', '']
]
    df = pd.DataFrame(arr)
    index=0

    # Now df is a Pandas DataFrame with the irregular structure preserved
    print(df)

    excel_file_path = "Outputs/Multipliers/Report.xlsx"
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active
        sheet.delete_rows(1, sheet.max_row)  # Clear all rows in the sheet

    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active 

    designs_directory = "Outputs/Multipliers/"
    for  design in os.listdir(designs_directory):
        if os.path.isdir(os.path.join(designs_directory, design)):
            process_design(index, design, sheet)
            index+=1
    for column_letter in "ABCDEF":
        sheet.column_dimensions[column_letter].width = 15

    for row in df.iterrows():
        sheet.append(list(row[1]))

    
    workbook.save(excel_file_path)
