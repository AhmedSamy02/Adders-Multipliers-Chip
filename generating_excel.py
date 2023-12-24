import re
import openpyxl
import numpy as np


def extract_min_timing_slack(report):
    # Extract slack value using regular expression
    slack_match = re.search(r'slack \(MET\)\s+([\d.-]+)', report)
    if slack_match:
        return float(slack_match.group(1))
    else:
        return None


def extract_timing_slack(report):
    # Extract slack value using regular expression
    slack_match = re.search(r'slack \(MET\)\s+([\d.-]+)', report)
    if slack_match:
        return float(slack_match.group(1))
    else:
        return None


def extract_power(report):
    # Extract slack value using regular expression
    match = re.search(r'Total\s+([\d.]+)\s+uW', report)
    slack_match = re.search(r'slack \(MET\)\s+([\d.-]+)', report)
    if slack_match:
        return float(slack_match.group(1))
    else:
        return None


def add_slack_to_excel(input_values, excel_file):

    # Add a new row with the slack value
    sheet.append(input_values)

    # Save the changes to the Excel file
    workbook.save(excel_file)


if __name__ == "__main__":
    arr = np.array([['', 'Verilog (*)', 'Multiplier Tree', 'Sequential Multiplier', 'Original Booth', 'Float Multiplier'],
                    ['Min Delay'], ['Max Delay'], ['Clock'],
                    ['Total Power'], ['Area'], ['Utilization'], ['# of Latches']])
    design = 'SimpleMultiplier'
    # Specify the path to your report file
    report_file_path = "Phase 2 reports/Simple Multiplier/"
    excel_file_path = "Phase 2 reports /reports.xlsx"
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
        # Select the active sheet
        sheet = workbook.active
        # Clear all rows in the sheet
        sheet.delete_rows(1, sheet.max_row)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    try:
        # Read the report from the file
        with open(report_file_path+design+'_synth_timing.rpt', 'r') as report_file:
            report_string = report_file.read()
        # Extract slack value from the report
        slack_value = extract_timing_slack(report_string)
        if slack_value is not None:
            # Add slack value to Excel sheet
            # Change the filename as needed
            add_slack_to_excel(slack_value, excel_file_path)
            print(f"Slack value added to {excel_file_path}")
        else:
            print("Error during extracting timing")
    except FileNotFoundError:
        print(f"Error: Report file not found at {report_file_path}")
    try:
        # Read the report from the file
        with open(report_file_path+design+'_synth_min_timing.rpt', 'r') as report_file:
            report_string = report_file.read()
        # Extract slack value from the report
        slack_value = extract_min_timing_slack(report_string)
        if slack_value is not None:
            # Add slack value to Excel sheet
            # Change the filename as needed
            add_slack_to_excel(slack_value, excel_file_path)
            print(f"Slack value added to {excel_file_path}")
        else:
            print("Error during extracting timing")
    except FileNotFoundError:
        print(f"Error: Report file not found at {report_file_path}")
